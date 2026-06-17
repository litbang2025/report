import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from io import BytesIO
import base64
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm, mm
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, 
    PageBreak, Image, KeepTogether
)
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import Font as XlFont, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Page config
st.set_page_config(
    page_title="DapurLab - Analisis Nilai Siswa",
    layout="wide",
    initial_sidebar_state="auto",
    page_icon="📊"
)

# ============ CSS ============
st.markdown("""
<style>
    * { font-family: 'Segoe UI', system-ui, -apple-system, sans-serif; }
    
    .main-header {
        background: linear-gradient(135deg, #1e3a5f, #2563eb);
        padding: 25px 30px; border-radius: 15px;
        color: white; margin-bottom: 25px;
    }
    .main-header h1 { margin: 0; font-size: 1.8rem; font-weight: 700; color: white !important; }
    .main-header p { margin: 5px 0 0 0; opacity: 0.9; font-size: 0.9rem; color: #cbd5e1 !important; }
    
    .stat-card {
        background: white; border-radius: 12px; padding: 20px;
        box-shadow: 0 1px 3px rgba(0,0,0,0.1); border: 1px solid #e5e7eb; margin-bottom: 15px;
    }
    .stat-card.primary { background: linear-gradient(135deg, #2563eb, #1d4ed8); color: white; border: none; }
    .stat-card.success { background: linear-gradient(135deg, #10b981, #059669); color: white; border: none; }
    .stat-label { font-size: 0.8rem; text-transform: uppercase; letter-spacing: 0.5px; opacity: 0.8; }
    .stat-value { font-size: 2rem; font-weight: 700; margin: 5px 0; }
    
    .subject-row {
        display: flex; align-items: center; padding: 10px 15px;
        border-bottom: 1px solid #f3f4f6; gap: 10px; flex-wrap: wrap;
    }
    .subject-row:hover { background: #f9fafb; }
    .subject-name { flex: 2; font-weight: 500; min-width: 140px; font-size: 0.9rem; }
    .subject-scores { display: flex; gap: 15px; align-items: center; flex: 1; min-width: 120px; }
    .subject-score { font-weight: 700; font-size: 1rem; min-width: 35px; text-align: center; }
    .subject-bar { flex: 3; height: 8px; background: #e5e7eb; border-radius: 4px; overflow: hidden; min-width: 80px; }
    .subject-bar-fill { height: 100%; border-radius: 4px; transition: width 0.5s; }
    
    .badge { display: inline-block; padding: 3px 10px; border-radius: 12px; font-size: 0.75rem; font-weight: 600; min-width: 70px; text-align: center; white-space: nowrap; }
    .badge-success { background: #dcfce7; color: #166534; }
    .badge-warning { background: #fef3c7; color: #92400e; }
    .badge-danger { background: #fee2e2; color: #991b1b; }
    .badge-info { background: #dbeafe; color: #1e40af; }
    
    .trend-up { color: #10b981; font-weight: 700; font-size: 0.8rem; }
    .trend-down { color: #ef4444; font-weight: 700; font-size: 0.8rem; }
    .trend-stable { color: #64748b; font-weight: 700; font-size: 0.8rem; }
    
    .section-title {
        font-size: 1.2rem; font-weight: 600; color: #1e293b;
        margin: 20px 0 10px 0; padding-bottom: 8px;
        border-bottom: 2px solid #e5e7eb;
    }
    
    .stButton > button {
        background: linear-gradient(135deg, #2563eb, #1d4ed8);
        color: white; border: none; border-radius: 10px;
        padding: 12px 30px; font-weight: 600; font-size: 1rem;
        width: 100%; cursor: pointer; transition: all 0.3s;
    }
    .stButton > button:hover { transform: translateY(-2px); box-shadow: 0 5px 15px rgba(37,99,235,0.3); }
    
    @media (max-width: 768px) {
        .main-header { padding: 15px; }
        .main-header h1 { font-size: 1.3rem; }
        .stat-value { font-size: 1.5rem; }
        .subject-name { min-width: 100px; font-size: 0.8rem; }
    }
    
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
</style>
""", unsafe_allow_html=True)


# ============ FUNCTIONS ============

def create_template_excel():
    """Create template Excel"""
    wb = openpyxl.Workbook()
    
    header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
    header_font = XlFont(color="FFFFFF", bold=True, size=11, name="Arial")
    example_fill = PatternFill(start_color="f0f9ff", end_color="f0f9ff", fill_type="solid")
    example_font = XlFont(italic=True, color="94a3b8", size=10, name="Arial")
    border = Border(
        left=Side(style='thin', color='d1d5db'),
        right=Side(style='thin', color='d1d5db'),
        top=Side(style='thin', color='d1d5db'),
        bottom=Side(style='thin', color='d1d5db')
    )
    section_fill = PatternFill(start_color="dbeafe", end_color="dbeafe", fill_type="solid")
    
    # Sheet PETUNJUK
    ws = wb.active
    ws.title = "PETUNJUK"
    ws.merge_cells('A1:H1')
    ws['A1'] = "PETUNJUK PENGISIAN DATA - DapurLab Analytics"
    ws['A1'].font = XlFont(bold=True, size=16, color="1e3a5f", name="Arial")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40
    
    instructions = [
        ("PENTING!", "Baris 1 adalah HEADER. JANGAN DIHAPUS!"),
        ("Kolom A", "NISN - WAJIB diisi (bisa angka atau teks)"),
        ("Kolom B", "Nama Siswa - WAJIB diisi"),
        ("Kolom C dst", "NILAI MAPEL - Bisa ditambah kolom baru"),
        ("", ""),
        ("ATURAN:", "Nilai 0-100, jumlah mapel TIDAK TERBATAS, nama mapel BOLEH disesuaikan"),
        ("TIPS:", "Sheet SEMESTER_2 opsional untuk perbandingan"),
    ]
    
    row = 3
    for label, text in instructions:
        if label in ["PENTING!", "ATURAN:", "TIPS:"]:
            ws.merge_cells(f'A{row}:H{row}')
            ws[f'A{row}'] = f"📌 {label}"
            ws[f'A{row}'].font = XlFont(bold=True, size=12, color="1e40af", name="Arial")
            ws[f'A{row}'].fill = section_fill
            row += 1
            if text:
                ws.merge_cells(f'A{row}:H{row}')
                ws[f'A{row}'] = text
                ws[f'A{row}'].font = XlFont(size=10, name="Arial")
                row += 1
        elif label:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = XlFont(bold=True, size=10, name="Arial")
            ws.merge_cells(f'B{row}:H{row}')
            ws[f'B{row}'] = text
            ws[f'B{row}'].font = XlFont(size=10, name="Arial")
            row += 1
        else:
            row += 1
    
    ws.column_dimensions['A'].width = 12
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 20
    
    # Sheet SEMESTER
    default_subjects = [
        "Adab dan Akhlak", "Al-Qur'an", "Aqidah", "Bahasa Arab",
        "Bahasa Indonesia", "Bahasa Inggris", "Fiqih", "IPA",
        "IPS", "Matematika", "PJOK", "PLBJ",
        "Pendidikan Pancasila", "Praktek Ibadah", "Seni Budaya", "Siroh"
    ]
    
    for sheet_name in ["SEMESTER_1", "SEMESTER_2"]:
        ws_data = wb.create_sheet(sheet_name)
        headers = ["NISN", "Nama Siswa"] + default_subjects
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        example_data = [1234567890, "CONTOH: Ahmad Fauzi"] + [85] * len(default_subjects)
        for col_idx, value in enumerate(example_data, 1):
            cell = ws_data.cell(row=3, column=col_idx, value=value)
            cell.fill = example_fill
            cell.font = example_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx in range(4, 54):
            for col_idx in range(1, len(headers) + 1):
                cell = ws_data.cell(row=row_idx, column=col_idx)
                cell.border = Border(
                    left=Side(style='hair', color='e5e7eb'),
                    right=Side(style='hair', color='e5e7eb'),
                    top=Side(style='hair', color='e5e7eb'),
                    bottom=Side(style='hair', color='e5e7eb')
                )
                cell.alignment = Alignment(horizontal='center')
        
        ws_data.column_dimensions['A'].width = 15
        ws_data.column_dimensions['B'].width = 30
        for col in range(3, len(headers) + 1):
            ws_data.column_dimensions[get_column_letter(col)].width = 18
        ws_data.freeze_panes = 'C2'
        ws_data.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def clean_dataframe(df):
    """Clean dataframe"""
    df = df.dropna(how='all')
    df = df.dropna(axis=1, how='all')
    
    name_col = None
    for col in df.columns:
        if 'nama' in str(col).lower():
            name_col = col
            break
    if name_col is None:
        name_col = df.columns[1] if len(df.columns) > 1 else df.columns[0]
    
    df = df.dropna(subset=[name_col])
    df = df[df[name_col].astype(str).str.strip() != '']
    df = df[~df[name_col].astype(str).str.contains('CONTOH|Contoh|contoh', na=False)]
    
    for col in df.columns:
        if col != name_col and 'nisn' not in str(col).lower():
            df[col] = pd.to_numeric(df[col], errors='coerce')
    
    return df, name_col


def detect_subjects(df):
    """Detect subject columns"""
    exclude = ['nisn', 'nama', 'no', 'id', 'induk']
    subjects = []
    for col in df.columns:
        col_lower = str(col).lower().strip()
        if not any(kw in col_lower for kw in exclude):
            subjects.append(col)
    return subjects


def get_grade(score, kkm=80):
    """Get grade"""
    if pd.isna(score): return "N/A", "badge-info", "Tidak ada nilai"
    if score >= 95: return "A+", "badge-success", "Sangat Istimewa"
    if score >= 90: return "A", "badge-success", "Istimewa"
    if score >= 85: return "B+", "badge-info", "Sangat Baik"
    if score >= kkm: return "B", "badge-info", "Baik"
    if score >= 70: return "C+", "badge-warning", "Cukup"
    if score >= 60: return "C", "badge-warning", "Kurang"
    return "D", "badge-danger", "Perlu Perbaikan"


def create_bar_chart_image(scores, subjects, kkm, title):
    """Create bar chart as image for PDF"""
    fig, ax = plt.subplots(figsize=(7, 3.5))
    
    colors_bar = []
    for s in scores:
        if pd.isna(s):
            colors_bar.append('#94a3b8')
        elif s >= 90:
            colors_bar.append('#10b981')
        elif s >= kkm:
            colors_bar.append('#3b82f6')
        elif s >= 70:
            colors_bar.append('#f59e0b')
        else:
            colors_bar.append('#ef4444')
    
    x_pos = range(len(subjects))
    bars = ax.bar(x_pos, scores, color=colors_bar, edgecolor='white', linewidth=0.5)
    
    # Add value labels
    for i, (bar, score) in enumerate(zip(bars, scores)):
        if not pd.isna(score):
            ax.text(bar.get_x() + bar.get_width()/2., bar.get_height() + 1,
                    f'{score:.0f}', ha='center', va='bottom', fontsize=7, fontweight='bold')
    
    # KKM line
    ax.axhline(y=kkm, color='red', linestyle='--', linewidth=1, label=f'KKM = {kkm}')
    
    ax.set_xticks(x_pos)
    ax.set_xticklabels(subjects, rotation=45, ha='right', fontsize=7)
    ax.set_ylim(0, 105)
    ax.set_ylabel('Nilai', fontsize=9)
    ax.set_title(title, fontsize=11, fontweight='bold', color='#1e3a5f')
    ax.legend(fontsize=8, loc='upper right')
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.grid(axis='y', alpha=0.3)
    
    plt.tight_layout()
    
    buf = BytesIO()
    plt.savefig(buf, format='png', dpi=150, bbox_inches='tight', facecolor='white')
    plt.close()
    buf.seek(0)
    return buf

def create_professional_pdf(nama, nisn, subjects, scores_s1, scores_s2, 
                            avg_s1, avg_s2, kkm, items, has_s2):
    """PROFESSIONAL PDF REPORT - A4 PORTRAIT (semua mapel ditampilkan)"""
    buffer = BytesIO()
    
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=1.8*cm,
        rightMargin=1.8*cm,
        topMargin=1.2*cm,
        bottomMargin=1.5*cm
    )
    
    styles = getSampleStyleSheet()
    story = []
    
    # Colors
    navy = colors.HexColor('#1e3a5f')
    blue = colors.HexColor('#2563eb')
    green = colors.HexColor('#10b981')
    red = colors.HexColor('#ef4444')
    gray = colors.HexColor('#64748b')
    dark = colors.HexColor('#1e293b')
    light_bg = colors.HexColor('#f8fafc')
    white = colors.white
    
    # Styles
    style_kop_title = ParagraphStyle(
        'KopTitle', fontName='Helvetica-Bold', fontSize=18,
        textColor=navy, alignment=TA_CENTER, spaceAfter=4, leading=22
    )
    style_kop_sub = ParagraphStyle(
        'KopSub', fontName='Helvetica', fontSize=8,
        textColor=gray, alignment=TA_CENTER, spaceAfter=2, leading=10
    )
    style_judul = ParagraphStyle(
        'Judul', fontName='Helvetica-Bold', fontSize=14,
        textColor=navy, alignment=TA_CENTER, spaceAfter=6, spaceBefore=10, leading=18
    )
    style_section = ParagraphStyle(
        'Section', fontName='Helvetica-Bold', fontSize=11,
        textColor=navy, spaceAfter=6, spaceBefore=10, leading=15
    )
    style_subsection = ParagraphStyle(
        'SubSection', fontName='Helvetica-Bold', fontSize=9,
        textColor=blue, spaceAfter=4, spaceBefore=6, leading=13
    )
    style_normal = ParagraphStyle(
        'Normal', fontName='Helvetica', fontSize=8.5,
        textColor=dark, leading=12, alignment=TA_JUSTIFY
    )
    style_small = ParagraphStyle(
        'Small', fontName='Helvetica', fontSize=7,
        textColor=gray, leading=9
    )
    style_center = ParagraphStyle(
        'Center', fontName='Helvetica', fontSize=8,
        alignment=TA_CENTER, leading=10
    )
    
    # ===== KOP SURAT =====
    kop_rows = [
        [Paragraph("🧪 DAPURLAB", style_kop_title)],
        [Paragraph("EDUCATION ANALYTICS PLATFORM", ParagraphStyle(
            'KopSub2', fontName='Helvetica-Bold', fontSize=10,
            textColor=blue, alignment=TA_CENTER, spaceAfter=2, leading=14))],
        [Paragraph("Munjul-Cipayung Jakarta Timur | Telp: 085640375704 | Email: kreatif.appmobile@gmail.com", style_kop_sub)],
    ]
    
    kop_table = Table(kop_rows, colWidths=[doc.width])
    kop_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 1),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
    ]))
    story.append(kop_table)
    
    # Double line separator
    story.append(Spacer(1, 0.15*cm))
    line_table = Table([['']], colWidths=[doc.width], rowHeights=[2])
    line_table.setStyle(TableStyle([
        ('LINEBELOW', (0, 0), (-1, 0), 2, blue),
    ]))
    story.append(line_table)
    story.append(Spacer(1, 0.1*cm))
    line_table2 = Table([['']], colWidths=[doc.width], rowHeights=[1])
    line_table2.setStyle(TableStyle([
        ('LINEBELOW', (0, 0), (-1, 0), 1, navy),
    ]))
    story.append(line_table2)
    story.append(Spacer(1, 0.6*cm))
    
    # ===== JUDUL =====
    story.append(Paragraph("LAPORAN HASIL ANALISIS AKADEMIK SISWA", style_judul))
    story.append(Paragraph(
        f"Tahun Ajaran 2025/2026 | Semester Ganjil{' & Genap' if has_s2 else ''}",
        ParagraphStyle('SubJudul', fontName='Helvetica', fontSize=9,
                      textColor=gray, alignment=TA_CENTER, spaceAfter=8)
    ))
    story.append(Spacer(1, 0.3*cm))
    
    # ===== IDENTITAS =====
    story.append(Paragraph("A. IDENTITAS SISWA", style_section))
    
    nisn_str = str(nisn)
    try:
        nisn_str = str(int(float(nisn)))
    except:
        pass
    
    # Hitung status kelulusan
    valid_s1 = [s for s in scores_s1 if not pd.isna(s)]
    total_s1 = len(valid_s1)
    lulus_s1 = sum(1 for s in valid_s1 if s >= kkm)
    remidi_count = total_s1 - lulus_s1
    if remidi_count == 0:
        status_lulus = "✅ Semua mapel LULUS"
    else:
        status_lulus = f"⚠️ {remidi_count} mapel perlu perbaikan (REMIDI)"
    
    id_data = [
        ['Nama Lengkap', ':', str(nama)],
        ['NISN', ':', nisn_str],
        ['Tanggal Laporan', ':', datetime.now().strftime('%d %B %Y')],
        ['KKM (Standar Kelulusan)', ':', str(kkm)],
        ['Status Kelulusan', ':', status_lulus],
        ['Jumlah Mata Pelajaran', ':', str(len(subjects))],
    ]
    
    id_table = Table(id_data, colWidths=[4.5*cm, 0.6*cm, 11*cm])
    id_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TEXTCOLOR', (0, 0), (0, -1), navy),
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LINEBELOW', (0, 0), (-1, -1), 0.3, colors.HexColor('#e5e7eb')),
    ]))
    story.append(id_table)
    story.append(Spacer(1, 0.5*cm))
    
    # ===== FITUR BARU: RINGKASAN UNTUK ORANG TUA =====
    story.append(Paragraph("👀 YANG PERLU DIPERHATIKAN", style_section))
    
    # Ambil 3 terendah dan 3 tertinggi
    sorted_items = sorted(items, key=lambda x: x['s1'])
    lowest_3 = sorted_items[:3]
    highest_3 = sorted_items[-3:][::-1] if len(sorted_items) >= 3 else sorted_items[::-1]
    
    if remidi_count > 0:
        intro = f"Ananda <b>{nama}</b> memiliki <b>{remidi_count} mata pelajaran</b> yang masih di bawah standar kelulusan ({kkm}). "
        intro += "Perhatikan mata pelajaran berikut:"
        story.append(Paragraph(intro, style_normal))
        story.append(Spacer(1, 0.2*cm))
        for i, item in enumerate(lowest_3, 1):
            rek = get_recommendation(item['subj'], item['s1'], kkm)
            story.append(Paragraph(
                f"{i}. <b>{item['subj']}</b> (Nilai: {item['s1']:.0f}) – {rek}",
                style_normal
            ))
    else:
        story.append(Paragraph(
            f"🎉 Selamat! Ananda <b>{nama}</b> telah mencapai standar kelulusan di semua mata pelajaran. "
            "Pertahankan prestasi ini dan terus tingkatkan kemampuan.",
            style_normal
        ))
        story.append(Spacer(1, 0.2*cm))
        story.append(Paragraph("<b>Pencapaian tertinggi:</b>", style_normal))
        for i, item in enumerate(highest_3, 1):
            story.append(Paragraph(
                f"{i}. <b>{item['subj']}</b> (Nilai: {item['s1']:.0f})",
                style_normal
            ))
    
    story.append(Spacer(1, 0.2*cm))
    story.append(Paragraph(
        "💡 <b>Tips untuk orang tua:</b> Diskusikan dengan anak tentang mata pelajaran yang perlu perbaikan. "
        "Buat jadwal belajar tambahan dan konsultasikan dengan guru jika diperlukan.",
        style_normal
    ))
    story.append(Spacer(1, 0.6*cm))
    
    # ===== RINGKASAN AKADEMIK =====
    story.append(Paragraph("B. RINGKASAN AKADEMIK", style_section))
    
    max_s1 = max(valid_s1) if valid_s1 else 0
    min_s1 = min(valid_s1) if valid_s1 else 0
    
    if avg_s1 >= 95: pred_s1 = "A+ (Sangat Istimewa)"
    elif avg_s1 >= 90: pred_s1 = "A (Istimewa)"
    elif avg_s1 >= 85: pred_s1 = "B+ (Sangat Baik)"
    elif avg_s1 >= kkm: pred_s1 = "B (Baik)"
    elif avg_s1 >= 70: pred_s1 = "C+ (Cukup)"
    else: pred_s1 = "D (Perlu Perbaikan)"
    
    pct_s1 = f'{(lulus_s1/total_s1*100):.0f}%' if total_s1 > 0 else '0%'
    
    if has_s2 and scores_s2:
        valid_s2 = [s for s in scores_s2 if s is not None and not pd.isna(s)]
        lulus_s2 = sum(1 for s in valid_s2 if s >= kkm)
        total_s2 = len(valid_s2)
        max_s2 = f'{max(valid_s2):.0f}' if valid_s2 else '-'
        min_s2 = f'{min(valid_s2):.0f}' if valid_s2 else '-'
        avg_s2_str = f'{avg_s2:.1f}' if avg_s2 else '-'
        pct_s2 = f'{(lulus_s2/total_s2*100):.0f}%' if total_s2 > 0 else '0%'
        pred_s2 = "A+" if avg_s2 and avg_s2 >= 95 else "A" if avg_s2 and avg_s2 >= 90 else "B+"
    else:
        lulus_s2 = '-'; total_s2 = '-'; max_s2 = '-'; min_s2 = '-'
        avg_s2_str = '-'; pct_s2 = '-'; pred_s2 = '-'
    
    ringkasan_data = [
        ['Indikator', 'Semester 1', 'Semester 2'],
        ['Rata-rata Nilai', f'{avg_s1:.1f}', avg_s2_str],
        ['Nilai Tertinggi', f'{max_s1:.0f}', str(max_s2)],
        ['Nilai Terendah', f'{min_s1:.0f}', str(min_s2)],
        ['Mapel Lulus KKM', f'{lulus_s1}/{total_s1} ({pct_s1})',
         f'{lulus_s2}/{total_s2} ({pct_s2})' if has_s2 else '-'],
        ['Predikat Akademik', pred_s1, pred_s2],
    ]
    
    ringkasan_table = Table(ringkasan_data, colWidths=[5*cm, 4.5*cm, 4.5*cm])
    ringkasan_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), navy),
        ('TEXTCOLOR', (0, 0), (-1, 0), white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [white, light_bg]),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TEXTCOLOR', (0, 1), (0, -1), navy),
    ]))
    story.append(ringkasan_table)
    story.append(Spacer(1, 0.6*cm))
    
    # ===== GRAFIK =====
    story.append(Paragraph("C. VISUALISASI NILAI", style_section))
    
    try:
        chart_buf = create_bar_chart_image(
            [s if not pd.isna(s) else 0 for s in scores_s1],
            subjects, kkm,
            f'Grafik Nilai Semester 1 - {nama}'
        )
        chart_img = Image(chart_buf, width=doc.width * 0.9, height=4*cm)
        story.append(chart_img)
    except Exception as e:
        story.append(Paragraph(f"(Grafik tidak dapat ditampilkan)", style_small))
    
    story.append(Spacer(1, 0.5*cm))
    
    # ===== HALAMAN 2 =====
    story.append(PageBreak())
    
    story.append(Paragraph("DAPURLAB EDUCATION ANALYTICS", ParagraphStyle(
        'KopHal2', fontName='Helvetica-Bold', fontSize=10,
        textColor=navy, alignment=TA_CENTER, spaceAfter=4)))
    story.append(Paragraph(f"Laporan: {nama} | NISN: {nisn_str}", style_small))
    story.append(Spacer(1, 0.3*cm))
    
    # ===== DETAIL LENGKAP (SEMUA MAPEL DITAMPILKAN) =====
    story.append(Paragraph("D. REKAPITULASI NILAI LENGKAP", style_section))
    story.append(Paragraph(
        f"Keterangan: KKM = {kkm} | Nilai ≥ {kkm} = LULUS | '-' berarti data tidak tersedia",
        style_small
    ))
    story.append(Spacer(1, 0.2*cm))
    
    detail_data = [['No', 'Mata Pelajaran', 'Nilai S1', 'Nilai S2', 'Trend', 'Status', 'Grade']]
    
    for i, subj in enumerate(subjects):
        s1 = scores_s1[i]
        s2 = scores_s2[i] if has_s2 and scores_s2 and i < len(scores_s2) else None
        
        # Jika nilai S1 kosong, tampilkan "-" dan status N/A
        if pd.isna(s1):
            s2_display = f'{s2:.0f}' if s2 is not None and not pd.isna(s2) else '-'
            detail_data.append([
                str(i+1), subj,
                '-',
                s2_display,
                '-',
                'N/A',
                '-'
            ])
            continue
        
        # Proses normal untuk nilai yang ada
        # Trend
        if s2 is not None and not pd.isna(s2):
            diff = s2 - s1
            if diff > 3:
                trend = f'▲ +{diff:.0f}'
            elif diff < -3:
                trend = f'▼ {diff:.0f}'
            else:
                trend = '➡️ Stabil'
        else:
            trend = '-'
        
        status = 'LULUS' if s1 >= kkm else 'REMIDI'
        grade, _, _ = get_grade(s1, kkm)
        
        detail_data.append([
            str(i+1), subj,
            f'{s1:.0f}',
            f'{s2:.0f}' if s2 is not None and not pd.isna(s2) else '-',
            trend, status, grade
        ])
    
    # Lebar kolom disesuaikan
    detail_table = Table(detail_data,
                        colWidths=[0.6*cm, 3.5*cm, 1.2*cm, 1.2*cm, 1.4*cm, 1.5*cm, 1.8*cm],
                        repeatRows=1)
    
    detail_style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), navy),
        ('TEXTCOLOR', (0, 0), (-1, 0), white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7.5),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#cbd5e1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [white, light_bg]),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]
    
    for i, row in enumerate(detail_data[1:], 1):
        if len(row) > 5 and row[5] == 'REMIDI':
            detail_style_cmds.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#fef2f2')))
            detail_style_cmds.append(('TEXTCOLOR', (5, i), (5, i), red))
        elif len(row) > 5 and row[5] == 'N/A':
            detail_style_cmds.append(('TEXTCOLOR', (0, i), (-1, i), gray))
    
    detail_table.setStyle(TableStyle(detail_style_cmds))
    story.append(detail_table)
    story.append(Spacer(1, 0.5*cm))
    
    # ===== KEKUATAN =====
    story.append(Paragraph("E. ANALISIS KEKUATAN & KELEMAHAN", style_section))
    story.append(Paragraph("Kekuatan Akademik (5 Nilai Tertinggi Semester 1)", style_subsection))
    
    sorted_asc = sorted(items, key=lambda x: x['s1'], reverse=True)
    
    str_data = [['No', 'Mata Pelajaran', 'Nilai S1', 'Nilai S2', 'Grade', 'Catatan']]
    for i, item in enumerate(sorted_asc[:5], 1):
        s2_str = f'{item["s2"]:.0f}' if item['s2'] is not None and not pd.isna(item['s2']) else '-'
        grade, _, label = get_grade(item['s1'], kkm)
        str_data.append([str(i), item['subj'], f'{item["s1"]:.0f}', s2_str,
                        f'{grade} ({label})', 'Pertahankan prestasi!'])
    
    str_table = Table(str_data, colWidths=[1*cm, 3.5*cm, 1.8*cm, 1.8*cm, 3*cm, 3*cm])
    str_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), green),
        ('TEXTCOLOR', (0, 0), (-1, 0), white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#cbd5e1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [white, colors.HexColor('#f0fdf4')]),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(str_table)
    story.append(Spacer(1, 0.3*cm))
    
    # ===== KELEMAHAN =====
    story.append(Paragraph("Kelemahan Akademik (5 Nilai Terendah Semester 1)", style_subsection))
    
    sorted_desc = sorted(items, key=lambda x: x['s1'])
    
    weak_data = [['No', 'Mata Pelajaran', 'Nilai S1', 'Nilai S2', 'Grade', 'Rekomendasi']]
    for i, item in enumerate(sorted_desc[:5], 1):
        s2_str = f'{item["s2"]:.0f}' if item['s2'] is not None and not pd.isna(item['s2']) else '-'
        grade, _, label = get_grade(item['s1'], kkm)
        rek = get_recommendation(item['subj'], item['s1'], kkm)
        weak_data.append([str(i), item['subj'], f'{item["s1"]:.0f}', s2_str,
                         f'{grade} ({label})', rek])
    
    weak_table = Table(weak_data, colWidths=[1*cm, 3.5*cm, 1.8*cm, 1.8*cm, 3*cm, 3*cm])
    weak_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), red),
        ('TEXTCOLOR', (0, 0), (-1, 0), white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#cbd5e1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [white, colors.HexColor('#fef2f2')]),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(weak_table)
    story.append(Spacer(1, 0.6*cm))
    
    # ===== CATATAN =====
    story.append(Paragraph("F. CATATAN & REKOMENDASI UNTUK ORANG TUA", style_section))
    
    weak_subjects_list = [item for item in items if item['s1'] < kkm]
    
    if weak_subjects_list:
        story.append(Paragraph(
            f"Berdasarkan hasil analisis, terdapat <b>{len(weak_subjects_list)} mata pelajaran</b> "
            f"yang nilainya masih di bawah Kriteria Ketuntasan Minimal (KKM = {kkm}). "
            f"Berikut adalah catatan dan rekomendasi untuk perbaikan:",
            style_normal
        ))
        story.append(Spacer(1, 0.2*cm))
        for i, item in enumerate(weak_subjects_list, 1):
            rek = get_recommendation(item['subj'], item['s1'], kkm)
            story.append(Paragraph(
                f"{i}. <b>{item['subj']}</b> (Nilai: {item['s1']:.0f}) – {rek}",
                style_normal
            ))
    else:
        story.append(Paragraph(
            "Alhamdulillah, seluruh nilai mata pelajaran telah mencapai KKM. "
            "Pertahankan prestasi yang sudah baik dan terus tingkatkan semangat belajar.",
            style_normal
        ))
    
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph("<b>Rekomendasi Umum:</b>", style_normal))
    story.append(Paragraph(
        "• Tetapkan jadwal belajar rutin di rumah (minimal 1-2 jam per hari)<br/>"
        "• Komunikasikan perkembangan belajar dengan wali kelas secara berkala<br/>"
        "• Libatkan siswa dalam kegiatan ekstrakurikuler untuk pengembangan minat dan bakat<br/>"
        "• Berikan apresiasi atas setiap pencapaian untuk meningkatkan motivasi belajar",
        style_normal
    ))
    
    story.append(Spacer(1, 0.6*cm))
    
    # ===== TANDA TANGAN =====
    story.append(Paragraph("_" * 60, ParagraphStyle('Line', alignment=TA_CENTER, fontSize=1)))
    story.append(Spacer(1, 0.3*cm))
    
    ttd_data = [
        ['', 'Jakarta, ' + datetime.now().strftime('%d %B %Y'), ''],
        ['Mengetahui,', '', 'Hormat Kami,'],
        ['Wali Kelas', '', 'Guru Kelas'],
        ['', '', ''],
        ['', '', ''],
        ['', '', ''],
        ['( ______________________ )', '', '( ______________________ )'],
        ['NIP. ..................................', '', '............'],
    ]
    
    ttd_table = Table(ttd_data, colWidths=[5*cm, 4*cm, 5*cm])
    ttd_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    story.append(ttd_table)
    
    story.append(Spacer(1, 0.5*cm))
    
    # Footer
    story.append(Paragraph("_" * 60, ParagraphStyle('FooterLine', alignment=TA_CENTER, fontSize=1)))
    story.append(Paragraph(
        f"Dokumen ini dibuat secara otomatis oleh DapurLab Education Analytics Platform | "
        f"© {datetime.now().year} | Laporan ini bersifat rahasia dan hanya ditujukan untuk orang tua/wali siswa.",
        ParagraphStyle('Footer', fontName='Helvetica', fontSize=6,
                      textColor=colors.HexColor('#94a3b8'), alignment=TA_CENTER, leading=8)
    ))
    
    doc.build(story)
    buffer.seek(0)
    return buffer

def get_recommendation(subj, score, kkm):
    """Memberikan rekomendasi spesifik berdasarkan mapel dan nilai"""
    if score >= kkm:
        if score >= 90:
            return "Pertahankan dengan pengayaan dan latihan lanjutan."
        else:
            return "Terus tingkatkan dengan latihan rutin."
    else:
        subj_lower = subj.lower()
        if "matematika" in subj_lower:
            return "Perbanyak latihan soal cerita dan operasi hitung setiap hari."
        elif "bahasa" in subj_lower and ("arab" in subj_lower or "inggris" in subj_lower):
            return "Hafalkan kosakata dan praktek percakapan sederhana."
        elif "bahasa" in subj_lower:
            return "Perbanyak membaca buku dan latihan menulis."
        elif "ipa" in subj_lower or "ilmu pengetahuan alam" in subj_lower:
            return "Pelajari konsep dasar dengan praktikum sederhana."
        elif "ips" in subj_lower or "ilmu pengetahuan sosial" in subj_lower:
            return "Baca buku sejarah/geografi dan diskusikan dengan teman."
        elif any(w in subj_lower for w in ["qur'an", "ibadah", "fiqih", "adab", "aqidah", "siroh"]):
            return "Perdalam pemahaman dengan diskusi dan baca terjemahan."
        elif "pjok" in subj_lower or "olahraga" in subj_lower:
            return "Tingkatkan latihan fisik dan koordinasi gerak."
        elif "seni" in subj_lower:
            return "Perbanyak latihan dan eksplorasi kreativitas."
        else:
            return "Tambah jam belajar dan ikuti program remedial."


# ============ SIDEBAR ============
with st.sidebar:
    st.markdown("## ⚙️ Pengaturan")
    
    with st.expander("🎯 KKM (Kriteria Ketuntasan Minimal)"):
        kkm = st.number_input("Nilai KKM", 0, 100, 80)
    
    st.markdown("---")
    st.markdown("## 📥 Template Excel")
    st.markdown("""
    <div style="background:#eff6ff; padding:12px; border-radius:8px; font-size:0.85rem;">
    <b>Format fleksibel:</b><br>
    • Kolom A: NISN<br>
    • Kolom B: Nama Siswa<br>
    • Kolom C+: Mapel (bebas)<br>
    <b>Mapel TIDAK TERBATAS!</b>
    </div>
    """, unsafe_allow_html=True)
    
    st.download_button(
        "📥 Download Template Excel",
        create_template_excel(),
        "Template_Nilai_DapurLab.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
    
    st.markdown("---")
    st.markdown("## 📤 Upload Data")
    uploaded_file = st.file_uploader("Pilih file Excel", type=["xlsx", "xls"])
    
    if uploaded_file:
        st.success("✅ File siap!")
    
    st.markdown("---")
    st.caption("© 2024 DapurLab | v6.0 Professional")


# ============ HEADER ============
st.markdown("""
<div class="main-header">
    <h1>🧪 DapurLab Analytics</h1>
    <p>Platform Analisis Kekuatan & Kelemahan Siswa • Laporan ke Orang Tua • Professional Grade</p>
</div>
""", unsafe_allow_html=True)


# ============ MAIN CONTENT ============
if uploaded_file is None:
    st.markdown("## 👋 Selamat Datang!")
    st.markdown("""
    <div style="background:#f8fafc; padding:30px; border-radius:15px; text-align:center; margin:20px 0;">
        <div style="font-size:4rem;">🧪</div>
        <h3>Platform Analisis Nilai Siswa Profesional</h3>
        <p style="color:#64748b;">Upload file Excel untuk memulai analisis lengkap</p>
    </div>
    """, unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown("""
        <div style="background:white; padding:25px; border-radius:12px; border:1px solid #e5e7eb; text-align:center;">
            <div style="font-size:2rem;">1️⃣</div>
            <h4>Download Template</h4>
            <p style="font-size:0.9rem; color:#64748b;">Format fleksibel, mapel tidak terbatas</p>
        </div>
        """, unsafe_allow_html=True)
    with col2:
        st.markdown("""
        <div style="background:white; padding:25px; border-radius:12px; border:1px solid #e5e7eb; text-align:center;">
            <div style="font-size:2rem;">2️⃣</div>
            <h4>Isi Data Nilai</h4>
            <p style="font-size:0.9rem; color:#64748b;">Isi NISN, Nama, dan nilai per mapel</p>
        </div>
        """, unsafe_allow_html=True)
    with col3:
        st.markdown("""
        <div style="background:white; padding:25px; border-radius:12px; border:1px solid #e5e7eb; text-align:center;">
            <div style="font-size:2rem;">3️⃣</div>
            <h4>Upload & Analisis</h4>
            <p style="font-size:0.9rem; color:#64748b;">Download laporan PDF profesional</p>
        </div>
        """, unsafe_allow_html=True)

else:
    try:
        xl = pd.ExcelFile(uploaded_file)
        sheet_names = xl.sheet_names
        
        df_s1 = pd.read_excel(uploaded_file, sheet_name='SEMESTER_1' if 'SEMESTER_1' in sheet_names else 0)
        df_s2 = pd.read_excel(uploaded_file, sheet_name='SEMESTER_2') if 'SEMESTER_2' in sheet_names else None
        
        df_s1, name_col = clean_dataframe(df_s1)
        if df_s2 is not None:
            df_s2, _ = clean_dataframe(df_s2)
        
        subjects = detect_subjects(df_s1)
        has_s2 = df_s2 is not None and len(df_s2) > 0
        
        if not subjects:
            st.error("❌ Tidak ada kolom mapel terdeteksi!")
            st.stop()
        
        st.success(f"✅ **{len(df_s1)} siswa** | **{len(subjects)} mapel** | S2: {'✅ Tersedia' if has_s2 else '❌ Tidak'}")
        
        # Pilih Siswa
        st.markdown("## 👨‍🎓 Pilih Siswa")
        col_s1, col_s2 = st.columns([1, 2])
        with col_s1:
            search = st.text_input("🔍 Cari:", placeholder="Ketik nama...")
        names = df_s1[name_col].dropna().unique()
        if search:
            names = [n for n in names if search.lower() in str(n).lower()]
        if len(names) == 0:
            st.warning("Tidak ditemukan.")
            st.stop()
        with col_s2:
            selected = st.selectbox("Nama Siswa:", names, label_visibility="collapsed")
        
        # Get data
        student = df_s1[df_s1[name_col] == selected].iloc[0]
        nisn_val = student.get('NISN', student.get('NIS', student.iloc[0]))
        
        scores_s1 = []
        for subj in subjects:
            val = student.get(subj, np.nan)
            scores_s1.append(float(val) if pd.notna(val) else np.nan)
        
        valid_s1 = [s for s in scores_s1 if not pd.isna(s)]
        avg_s1 = np.mean(valid_s1) if valid_s1 else 0
        
        scores_s2 = None
        avg_s2 = None
        if has_s2 and selected in df_s2[name_col].values:
            student_s2 = df_s2[df_s2[name_col] == selected].iloc[0]
            scores_s2 = []
            for subj in subjects:
                if subj in df_s2.columns:
                    val = student_s2.get(subj, np.nan)
                    scores_s2.append(float(val) if pd.notna(val) else np.nan)
                else:
                    scores_s2.append(np.nan)
            valid_s2 = [s for s in scores_s2 if not pd.isna(s)]
            avg_s2 = np.mean(valid_s2) if valid_s2 else None
        
        # Build items
        items = []
        for i, subj in enumerate(subjects):
            s1 = scores_s1[i]
            if pd.isna(s1): continue
            s2 = scores_s2[i] if scores_s2 else None
            items.append({'subj': subj, 's1': s1, 's2': s2})
        
        # Stats Cards
        st.markdown("## 📊 Ringkasan Akademik")
        c1, c2, c3, c4 = st.columns(4)
        grade, badge, label = get_grade(avg_s1, kkm)
        
        with c1:
            st.markdown(f"""
            <div class="stat-card primary">
                <div class="stat-label">Rata-rata S1</div>
                <div class="stat-value">{avg_s1:.1f}</div>
                <span class="badge {badge}">{grade} - {label}</span>
            </div>
            """, unsafe_allow_html=True)
        
        with c2:
            if avg_s2 is not None:
                grade2, badge2, label2 = get_grade(avg_s2, kkm)
                diff = avg_s2 - avg_s1
                trend = "📈" if diff > 3 else "📉" if diff < -3 else "➡️"
                st.markdown(f"""
                <div class="stat-card success">
                    <div class="stat-label">Rata-rata S2</div>
                    <div class="stat-value">{avg_s2:.1f}</div>
                    <span>{trend} {diff:+.1f} | {grade2}</span>
                </div>
                """, unsafe_allow_html=True)
            else:
                st.markdown("""
                <div class="stat-card" style="background:#f8fafc;">
                    <div class="stat-label">Semester 2</div>
                    <div class="stat-value" style="color:#94a3b8;">-</div>
                </div>
                """, unsafe_allow_html=True)
        
        with c3:
            max_val = max(valid_s1)
            max_subj = subjects[scores_s1.index(max_val)]
            st.markdown(f"""
            <div class="stat-card" style="border-left:4px solid #10b981;">
                <div class="stat-label">Nilai Tertinggi</div>
                <div class="stat-value" style="color:#1e293b;">{max_val:.0f}</div>
                <span style="font-size:0.85rem;">{max_subj}</span>
            </div>
            """, unsafe_allow_html=True)
        
        with c4:
            min_val = min(valid_s1)
            min_subj = subjects[scores_s1.index(min_val)]
            icon = "⚠️" if min_val < kkm else "✅"
            st.markdown(f"""
            <div class="stat-card" style="border-left:4px solid {'#ef4444' if min_val < kkm else '#10b981'};">
                <div class="stat-label">Nilai Terendah</div>
                <div class="stat-value" style="color:#1e293b;">{min_val:.0f} {icon}</div>
                <span style="font-size:0.85rem;">{min_subj}</span>
            </div>
            """, unsafe_allow_html=True)
        
        # Status
        lulus = sum(1 for s in valid_s1 if s >= kkm)
        cols = st.columns(4)
        cols[0].metric("✅ Lulus", f"{lulus}/{len(valid_s1)}")
        cols[1].metric("⚠️ Remidi", f"{len(valid_s1)-lulus}")
        cols[2].metric("🌟 Istimewa", f"{sum(1 for s in valid_s1 if s>=90)}")
        cols[3].metric("📊 %", f"{(lulus/len(valid_s1)*100):.0f}%")
        
        # Detail
        st.markdown("## 📋 Detail Nilai")
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            status_filter = st.multiselect("Filter:", ["Lulus", "Remidi", "Istimewa"],
                                          default=["Lulus", "Remidi", "Istimewa"])
        with col_f2:
            sort_by = st.selectbox("Urutkan:", ["Nilai Tertinggi", "Nilai Terendah", "Abjad"])
        
        filtered_items = []
        for item in items:
            if item['s1'] >= 90: status = "Istimewa"
            elif item['s1'] >= kkm: status = "Lulus"
            else: status = "Remidi"
            if status in status_filter:
                item['status'] = status
                filtered_items.append(item)
        
        if sort_by == "Nilai Tertinggi":
            filtered_items.sort(key=lambda x: x['s1'], reverse=True)
        elif sort_by == "Nilai Terendah":
            filtered_items.sort(key=lambda x: x['s1'])
        else:
            filtered_items.sort(key=lambda x: x['subj'])
        
        for item in filtered_items:
            if item['s1'] >= 90: bar_color = "#10b981"
            elif item['s1'] >= kkm: bar_color = "#3b82f6"
            elif item['s1'] >= 70: bar_color = "#f59e0b"
            else: bar_color = "#ef4444"
            
            s2_display = ""
            trend_html = ""
            if item['s2'] is not None and not pd.isna(item['s2']):
                diff = item['s2'] - item['s1']
                trend_icon = "📈" if diff > 3 else "📉" if diff < -3 else "➡️"
                trend_class = "trend-up" if diff > 3 else "trend-down" if diff < -3 else "trend-stable"
                s2_display = f'<span style="margin-left:10px;font-size:0.85rem;">S2: <b>{item["s2"]:.0f}</b></span>'
                trend_html = f'<span class="{trend_class}">{trend_icon} {diff:+.0f}</span>'
            
            grade, badge, _ = get_grade(item['s1'], kkm)
            
            st.markdown(f"""
            <div class="subject-row">
                <div class="subject-name">{item['subj']}</div>
                <div class="subject-scores">
                    <div class="subject-score" style="color:{bar_color}">S1: {item['s1']:.0f}</div>
                    {s2_display} {trend_html}
                </div>
                <div class="subject-bar">
                    <div class="subject-bar-fill" style="width:{item['s1']}%;background:{bar_color};"></div>
                </div>
                <span class="badge {badge}">{grade}</span>
                <span style="font-size:0.75rem;">{item['status']}</span>
            </div>
            """, unsafe_allow_html=True)
        
        # Visualisasi
        st.markdown("## 📈 Visualisasi")
        tab1, tab2 = st.tabs(["📊 Bar Chart", "🎯 Radar Chart"])
        
        valid_idx = [i for i, s in enumerate(scores_s1) if not pd.isna(s)]
        plot_subjects = [subjects[i] for i in valid_idx]
        plot_s1 = [scores_s1[i] for i in valid_idx]
        
        with tab1:
            colors_bar = ['#10b981' if s >= 90 else '#3b82f6' if s >= kkm else '#f59e0b' if s >= 70 else '#ef4444' for s in plot_s1]
            fig = go.Figure()
            fig.add_trace(go.Bar(x=plot_subjects, y=plot_s1, marker_color=colors_bar, name='S1',
                                text=[f'{s:.0f}' for s in plot_s1], textposition='outside'))
            if scores_s2:
                plot_s2 = [scores_s2[i] for i in valid_idx]
                fig.add_trace(go.Bar(x=plot_subjects, y=plot_s2, marker_color='rgba(16,185,129,0.7)', name='S2'))
            fig.add_hline(y=kkm, line_dash="dash", line_color="red", annotation_text=f"KKM={kkm}")
            fig.update_layout(height=500, template='plotly_white', xaxis_tickangle=-45, yaxis=dict(range=[0,105]))
            st.plotly_chart(fig, use_container_width=True)
        
        with tab2:
            fig = go.Figure()
            fig.add_trace(go.Scatterpolar(r=plot_s1, theta=plot_subjects, fill='toself', name='S1', line_color='#3b82f6'))
            if scores_s2:
                fig.add_trace(go.Scatterpolar(r=[scores_s2[i] for i in valid_idx], theta=plot_subjects, fill='toself', name='S2', line_color='#10b981'))
            fig.update_layout(polar=dict(radialaxis=dict(range=[0,100])), height=500)
            st.plotly_chart(fig, use_container_width=True)
        
        # Download PDF
        st.markdown("## 📄 Download Laporan PDF Profesional")
        st.markdown("""
        <div style="background:#eff6ff; padding:15px; border-radius:10px; margin-bottom:15px;">
            <b>📋 Isi Laporan:</b> Kop Surat • Identitas • Ringkasan • Grafik • 
            Rekapitulasi Lengkap • Kekuatan & Kelemahan • Catatan Orang Tua • Tanda Tangan
        </div>
        """, unsafe_allow_html=True)
        
        if st.button("📑 Buat Laporan PDF Profesional", use_container_width=True, type="primary"):
            try:
                with st.spinner("📄 Membuat laporan profesional..."):
                    pdf_buffer = create_professional_pdf(
                        selected, nisn_val, subjects, scores_s1, scores_s2,
                        avg_s1, avg_s2, kkm, filtered_items, has_s2
                    )
                    b64 = base64.b64encode(pdf_buffer.getvalue()).decode()
                    
                    st.markdown(f"""
                    <div style="text-align:center;padding:20px;background:#f0fdf4;border-radius:12px;border:2px dashed #10b981;margin-top:15px;">
                        <p style="color:#166534;font-weight:700;font-size:1.1rem;">✅ Laporan Profesional Siap!</p>
                        <a href="data:application/pdf;base64,{b64}" 
                           download="Laporan_Akademik_{str(selected).replace(' ','_')}.pdf"
                           style="text-decoration:none;">
                            <button style="background:#10b981;color:white;border:none;padding:12px 35px;
                                           border-radius:10px;font-size:1rem;cursor:pointer;font-weight:700;
                                           margin-top:10px;">
                                📥 Download Laporan PDF
                            </button>
                        </a>
                    </div>
                    """, unsafe_allow_html=True)
                    st.balloons()
            except Exception as e:
                st.error(f"❌ Error: {str(e)}")
                import traceback
                st.code(traceback.format_exc())
    
    except Exception as e:
        st.error(f"❌ Error: {str(e)}")
        import traceback
        st.code(traceback.format_exc())
